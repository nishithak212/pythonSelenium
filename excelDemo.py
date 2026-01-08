import openpyxl

#Location of excel file
book = openpyxl.load_workbook("C:\\Users\\nishi\\PycharmProjects\\PythonTesting\\pythonSelenium\\excelFiles\\pythonExcelDemo.xlsx")
sheet = book.active #to get control of active excel sheet
dict = {} #empty dictionary

#to get control of a cell in the excel sheet
cell = sheet.cell(row = 1, column = 2)

#to get value of the cell on which the control is currently
print(cell.value)

#to write data to the cell in the sheet
sheet.cell(row=2, column = 2).value = "Pooh"

#print the value written in the cell
print(sheet.cell(row=2, column = 2).value)

#to get max rows in the sheet
print(sheet.max_row)

#to get max column in the sheet
print(sheet.max_column)

#to print values using cell number
print(sheet['A2'].value)
print(sheet['B2'].value)

#for loop to print all the values in the active excel sheet
for i in range(1, sheet.max_row+1): #to get rows
    #to scan only until a specific row
    if sheet.cell(row=i, column=1).value == "Testcase2": #This gives data sepcific to this testcase
    
        for j in range(1, sheet.max_column+1): #to get columns
            #print(sheet.cell(row=i, column=j).value) #to print the data in row that matches "Testcase2"
            #To store data in a dictionary
            dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
print(dict) #print values stored in dict -- output --> {'Name': 'Testcase2', 'firstname': 'Tiger', 'lastname': 'Brave', 'gender': 'M'}





